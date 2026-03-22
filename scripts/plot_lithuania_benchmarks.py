#!/usr/bin/env python3
"""
Plot Lithuania retention benchmark series from the Excel workbook "Model" sheet.

Output files are named so slides and archives stay traceable to the source .xlsx.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np

try:
    import matplotlib.pyplot as plt
except ImportError as e:
    raise SystemExit(
        "matplotlib is required. Install with: pip install -r requirements-figures.txt"
    ) from e

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit(
        "openpyxl is required. Install with: pip install -r requirements-figures.txt"
    ) from e


# Model sheet layout: row 4 = first year, row 29 = end year (packaged workbook)
FIRST_ROW = 4
LAST_ROW = 29
YEAR_COL = 1  # column A

# Output filename stem (stable prefix for sorting and search)
BENCHMARK_PREFIX = "lithuania-retention-benchmark"

SCENARIOS = [
    {
        "key": "Baseline",
        "pop": 2,
        "work": 3,
        "old": 4,
        "dep": 5,
        "net": 6,
        "fiscal": 7,
        "color": "#1f77b4",
    },
    {
        "key": "Moderate",
        "pop": 8,
        "work": 9,
        "old": 10,
        "dep": 11,
        "net": 12,
        "fiscal": 13,
        "color": "#ff7f0e",
    },
    {
        "key": "Strong",
        "pop": 14,
        "work": 15,
        "old": 16,
        "dep": 17,
        "net": 18,
        "fiscal": 19,
        "color": "#2ca02c",
    },
    {
        "key": "Stress",
        "pop": 20,
        "work": 21,
        "old": 22,
        "dep": 23,
        "net": 24,
        "fiscal": 25,
        "color": "#d62728",
    },
]

def workbook_slug(path: Path) -> str:
    """File-safe id from workbook basename, e.g. final_lithuania_model (1).xlsx -> final_lithuania_model_1."""
    stem = path.stem
    s = stem.replace("(", "_").replace(")", "").replace(" ", "_")
    s = re.sub(r"[^0-9A-Za-z._-]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "workbook"


def provenance_caption(workbook_path: Path) -> str:
    name = workbook_path.name
    return (
        f"Source file: {name}  |  Location: Model sheet  |  Series: rows {FIRST_ROW}–{LAST_ROW}  |  "
        f"Values: Excel cached (press Recalculate before export if needed)"
    )


def add_figure_footer(fig, caption: str) -> None:
    fig.subplots_adjust(bottom=0.18)
    fig.text(
        0.5,
        0.06,
        caption,
        ha="center",
        va="top",
        fontsize=8,
        color="#333333",
        transform=fig.transFigure,
    )


def outfile_name(slug: str, order: str, series_slug: str) -> str:
    """e.g. lithuania-retention-benchmark__final_lithuania_model_1__01_population-total_Model-sheet.png"""
    return f"{BENCHMARK_PREFIX}__{slug}__{order}_{series_slug}_Model-sheet.png"


def read_series(ws, row_start: int, row_end: int, col: int) -> np.ndarray:
    vals = []
    for r in range(row_start, row_end + 1):
        c = ws.cell(row=r, column=col)
        v = c.value
        if v is None:
            vals.append(np.nan)
        else:
            vals.append(float(v))
    return np.asarray(vals, dtype=float)


def read_years(ws) -> np.ndarray:
    return read_series(ws, FIRST_ROW, LAST_ROW, YEAR_COL)


def ensure_style():
    plt.rcParams.update(
        {
            "figure.figsize": (9, 5.75),
            "figure.dpi": 120,
            "savefig.dpi": 150,
            "font.size": 10,
            "axes.titlesize": 12,
            "axes.labelsize": 10,
            "legend.fontsize": 9,
            "axes.grid": True,
            "grid.alpha": 0.25,
        }
    )


def plot_population(years, data, out_path: Path, title_main: str, footer: str):
    fig, ax = plt.subplots()
    for s in SCENARIOS:
        ax.plot(
            years,
            data[s["key"]]["pop"] / 1e6,
            label=s["key"],
            color=s["color"],
            linewidth=2,
        )
    ax.set_xlabel("Year")
    ax.set_ylabel("Population (millions)")
    ax.set_title(title_main)
    ax.legend(loc="best", framealpha=0.95)
    add_figure_footer(fig, footer)
    fig.tight_layout(rect=[0, 0.14, 1, 1])
    fig.savefig(out_path)
    plt.close(fig)


def plot_working_age(years, data, out_path: Path, title_main: str, footer: str):
    fig, ax = plt.subplots()
    for s in SCENARIOS:
        ax.plot(
            years,
            data[s["key"]]["work"] / 1e6,
            label=s["key"],
            color=s["color"],
            linewidth=2,
        )
    ax.set_xlabel("Year")
    ax.set_ylabel("Working-age population (millions)")
    ax.set_title(title_main)
    ax.legend(loc="best", framealpha=0.95)
    add_figure_footer(fig, footer)
    fig.tight_layout(rect=[0, 0.14, 1, 1])
    fig.savefig(out_path)
    plt.close(fig)


def plot_dependency_ratio(years, data, out_path: Path, title_main: str, footer: str):
    fig, ax = plt.subplots()
    for s in SCENARIOS:
        ax.plot(
            years,
            data[s["key"]]["dep"] * 100,
            label=s["key"],
            color=s["color"],
            linewidth=2,
        )
    ax.set_xlabel("Year")
    ax.set_ylabel("Old-age dependency ratio (% of working-age)")
    ax.set_title(title_main)
    ax.legend(loc="best", framealpha=0.95)
    add_figure_footer(fig, footer)
    fig.tight_layout(rect=[0, 0.14, 1, 1])
    fig.savefig(out_path)
    plt.close(fig)


def plot_net_retained(years, data, out_path: Path, title_main: str, footer: str):
    fig, ax = plt.subplots()
    for s in SCENARIOS:
        ax.plot(
            years,
            data[s["key"]]["net"],
            label=s["key"],
            color=s["color"],
            linewidth=2,
        )
    ax.set_xlabel("Year")
    ax.set_ylabel("Net retained contributors / year")
    ax.set_title(title_main)
    ax.legend(loc="best", framealpha=0.95)
    add_figure_footer(fig, footer)
    fig.tight_layout(rect=[0, 0.14, 1, 1])
    fig.savefig(out_path)
    plt.close(fig)


def plot_fiscal_index(years, data, out_path: Path, title_main: str, footer: str):
    fig, ax = plt.subplots()
    for s in SCENARIOS:
        y = data[s["key"]]["fiscal"]
        ax.plot(years, y / 1e9, label=s["key"], color=s["color"], linewidth=2)
    ax.set_xlabel("Year")
    ax.set_ylabel("Stylised fiscal index (€ billions, workbook scale)")
    ax.set_title(title_main)
    ax.legend(loc="best", framealpha=0.95)
    add_figure_footer(fig, footer)
    fig.tight_layout(rect=[0, 0.14, 1, 1])
    fig.savefig(out_path)
    plt.close(fig)


def plot_population_delta_vs_baseline(years, data, out_path: Path, title_main: str, footer: str):
    base = data["Baseline"]["pop"]
    fig, ax = plt.subplots()
    for s in SCENARIOS:
        if s["key"] == "Baseline":
            continue
        delta = (data[s["key"]]["pop"] - base) / 1e3
        ax.plot(years, delta, label=f'{s["key"]} vs Baseline', color=s["color"], linewidth=2)
    ax.axhline(0, color="gray", linestyle="--", linewidth=0.8)
    ax.set_xlabel("Year")
    ax.set_ylabel("Population difference vs Baseline (thousands)")
    ax.set_title(title_main)
    ax.legend(loc="best", framealpha=0.95)
    add_figure_footer(fig, footer)
    fig.tight_layout(rect=[0, 0.14, 1, 1])
    fig.savefig(out_path)
    plt.close(fig)


def plot_combined_dashboard(years, data, out_path: Path, title_main: str, footer: str):
    fig, axes = plt.subplots(2, 2, figsize=(11, 8.2), dpi=120)
    for s in SCENARIOS:
        axes[0, 0].plot(
            years, data[s["key"]]["pop"] / 1e6, label=s["key"], color=s["color"], lw=1.8
        )
    axes[0, 0].set_title("Total population (millions)")
    axes[0, 0].set_xlabel("Year")
    axes[0, 0].legend(fontsize=8)

    for s in SCENARIOS:
        axes[0, 1].plot(
            years,
            data[s["key"]]["dep"] * 100,
            label=s["key"],
            color=s["color"],
            lw=1.8,
        )
    axes[0, 1].set_title("Dependency ratio (% of working-age)")
    axes[0, 1].set_xlabel("Year")

    base = data["Baseline"]["pop"]
    for s in SCENARIOS:
        if s["key"] == "Baseline":
            continue
        axes[1, 0].plot(
            years,
            (data[s["key"]]["pop"] - base) / 1e3,
            label=s["key"],
            color=s["color"],
            lw=1.8,
        )
    axes[1, 0].axhline(0, color="gray", ls="--", lw=0.7)
    axes[1, 0].set_title("Population vs Baseline (thousands)")
    axes[1, 0].set_xlabel("Year")
    axes[1, 0].legend(fontsize=8)

    for s in SCENARIOS:
        axes[1, 1].plot(
            years,
            data[s["key"]]["net"],
            label=s["key"],
            color=s["color"],
            lw=1.8,
        )
    axes[1, 1].set_title("Net retained / year")
    axes[1, 1].set_xlabel("Year")

    fig.suptitle(title_main, fontsize=13, y=1.02)
    add_figure_footer(fig, footer)
    fig.tight_layout(rect=[0, 0.12, 1, 0.98])
    fig.savefig(out_path, bbox_inches="tight")
    plt.close(fig)


def export_csv(years: np.ndarray, data: dict, path: Path, workbook_path: Path) -> Path:
    import csv

    header = ["year", "source_xlsx", "excel_sheet", "model_row_range"]
    for s in SCENARIOS:
        k = s["key"]
        header.extend(
            [
                f"{k}_pop",
                f"{k}_work",
                f"{k}_old",
                f"{k}_dependency_ratio",
                f"{k}_net_retained",
                f"{k}_fiscal_index",
            ]
        )
    meta = [workbook_path.name, "Model", f"rows_{FIRST_ROW}_{LAST_ROW}"]
    rows = []
    for i, yr in enumerate(years):
        row = [int(yr) if yr == int(yr) else yr, *meta]
        for s in SCENARIOS:
            k = s["key"]
            d = data[k]
            row.extend(
                [
                    d["pop"][i],
                    d["work"][i],
                    d["old"][i],
                    d["dep"][i],
                    d["net"][i],
                    d["fiscal"][i],
                ]
            )
        rows.append(row)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    return path


def load_model_data(workbook_path: Path):
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if "Model" in wb.sheetnames:
            ws = wb["Model"]
        else:
            ws = wb.worksheets[2]
    except IndexError as e:
        wb.close()
        raise SystemExit(
            "Workbook must contain a sheet named 'Model' or at least 3 tabs "
            "(Model as third sheet)."
        ) from e

    years = read_years(ws)
    data = {}
    for s in SCENARIOS:
        k = s["key"]
        data[k] = {
            "pop": read_series(ws, FIRST_ROW, LAST_ROW, s["pop"]),
            "work": read_series(ws, FIRST_ROW, LAST_ROW, s["work"]),
            "old": read_series(ws, FIRST_ROW, LAST_ROW, s["old"]),
            "dep": read_series(ws, FIRST_ROW, LAST_ROW, s["dep"]),
            "net": read_series(ws, FIRST_ROW, LAST_ROW, s["net"]),
            "fiscal": read_series(ws, FIRST_ROW, LAST_ROW, s["fiscal"]),
        }
    wb.close()
    return years, data


def resolve_repo_root() -> Path:
    return Path(__file__).resolve().parent.parent


def default_workbook(repo_root: Path) -> Path:
    candidates = [
        repo_root / "workbooks" / "final_lithuania_model (1).xlsx",
        repo_root / "workbooks" / "final_lithuania_model (2).xlsx",
        repo_root / "final_lithuania_model (1).xlsx",
        repo_root / "final_lithuania_model (2).xlsx",
    ]
    for p in candidates:
        if p.is_file():
            return p
    raise SystemExit(
        "No workbook found. Add .xlsx under workbooks/ or repo root, or pass --workbook PATH"
    )


def main():
    parser = argparse.ArgumentParser(
        description="Plot Lithuania benchmark charts from Excel Model sheet (named outputs per workbook)."
    )
    parser.add_argument("--workbook", type=Path, default=None, help="Path to .xlsx")
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output directory root (default: outputs/figures/<workbook-slug>/)",
    )
    parser.add_argument("--no-csv", action="store_true", help="Skip CSV export")
    args = parser.parse_args()

    repo_root = resolve_repo_root()
    wb_path = (args.workbook or default_workbook(repo_root)).resolve()
    if not wb_path.is_file():
        raise SystemExit(f"Workbook not found: {wb_path}")

    slug = workbook_slug(wb_path)
    out_dir = args.out or (repo_root / "outputs" / "figures" / slug)
    out_dir.mkdir(parents=True, exist_ok=True)

    footer = provenance_caption(wb_path)
    title_core = "Lithuania — scenario benchmark (Euro Challenge model)"

    ensure_style()
    years, data = load_model_data(wb_path)

    plot_population(
        years,
        data,
        out_dir / outfile_name(slug, "01", "population-total"),
        f"{title_core}: total population",
        footer,
    )
    plot_working_age(
        years,
        data,
        out_dir / outfile_name(slug, "02", "working-age-population"),
        f"{title_core}: working-age population",
        footer,
    )
    plot_dependency_ratio(
        years,
        data,
        out_dir / outfile_name(slug, "03", "old-age-dependency-ratio"),
        f"{title_core}: old-age dependency ratio",
        footer,
    )
    plot_net_retained(
        years,
        data,
        out_dir / outfile_name(slug, "04", "net-retained-contributors"),
        f"{title_core}: net retained contributors per year\n(Baseline column F = workbook constant)",
        footer,
    )
    plot_fiscal_index(
        years,
        data,
        out_dir / outfile_name(slug, "05", "fiscal-index-stylised"),
        f"{title_core}: stylised fiscal index from workbook",
        footer,
    )
    plot_population_delta_vs_baseline(
        years,
        data,
        out_dir / outfile_name(slug, "06", "population-delta-vs-baseline"),
        f"{title_core}: population relative to Baseline",
        footer,
    )
    plot_combined_dashboard(
        years,
        data,
        out_dir / outfile_name(slug, "07", "dashboard-2x2"),
        f"{title_core} — slide dashboard",
        footer,
    )

    csv_path = None
    if not args.no_csv:
        csv_name = f"{BENCHMARK_PREFIX}__{slug}__Model-sheet_time-series_wide.csv"
        csv_path = export_csv(years, data, out_dir / csv_name, wb_path)

    print(f"Workbook: {wb_path}")
    print(f"Workbook slug: {slug}")
    print(f"Outputs: {out_dir.resolve()}")
    for p in sorted(out_dir.glob("*.png")):
        print(f"  - {p.name}")
    if csv_path:
        print(f"CSV: {csv_path.resolve()}")


if __name__ == "__main__":
    main()
